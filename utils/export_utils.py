# c:\Users\mahmo\OneDrive\Documents\GitHub\hr_management_system\utils\export_utils.py
import openpyxl
from openpyxl.utils import get_column_letter
import logging
from typing import List, Dict, Any

logger = logging.getLogger(__name__)

def export_payroll_to_excel(data_list: List[Dict[str, Any]], filepath: str) -> bool:
    """
    Exports payroll data to an Excel file.
    Each dictionary in data_list represents a row.
    Keys of the first dictionary are used as headers.
    """
    if not data_list:
        logger.warning("No data provided for Excel export.")
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None: # Should not happen with a new workbook
        logger.error("Failed to get active sheet from new workbook.")
        return False
    ws.title = "Payroll Export"

    headers = list(data_list[0].keys())
    ws.append(headers)

    for row_data in data_list:
        row_values = [row_data.get(header, "") for header in headers]
        ws.append(row_values)
    
    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        if ws[column_letter + '1'].value:
             max_length = len(str(ws[column_letter + '1'].value))
        
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: # pragma: no cover
                pass # Ignore errors for empty cells or unexpected types during len check
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    try:
        wb.save(filepath)
        logger.info(f"Payroll data successfully exported to Excel: {filepath}")
        return True
    except Exception as e:
        logger.error(f"Error saving Excel file to {filepath}: {e}", exc_info=True)
        return False